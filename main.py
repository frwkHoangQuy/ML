import pandas as pd
import requests
from openpyxl import load_workbook
from openpyxl.styles import Alignment


def data_to_excel(data):
    # Create DataFrame
    df = pd.DataFrame(data)

    # Sort DataFrame by driver name (replace 'Lái xe' with the actual column name)
    df_sorted = df.sort_values(by='Lái xe')

    # Calculate total amount per driver
    df_sorted['Tổng tiền'] = df_sorted.groupby('Lái xe')['Thành tiền'].transform('sum')

    # Save to Excel
    output_file_path = 'output_excel_sorted.xlsx'
    df_sorted.to_excel(output_file_path, index=False)

    # Load the Excel file and get the active worksheet
    wb = load_workbook(output_file_path)
    ws = wb.active

    # Columns to exclude from merging
    exclude_columns = ['Ngày', 'Hạng mục thanh toán', 'Đơn giá', 'Số lượng', 'Thành tiền']
    exclude_columns_indices = [df_sorted.columns.get_loc(col) + 1 for col in
                               exclude_columns]  # +1 because Excel columns are 1-based

    # Function to merge cells in a column range
    def merge_column_range(start_row, end_row, column):
        if end_row > start_row:
            ws.merge_cells(start_row=start_row, start_column=column, end_row=end_row, end_column=column)
            cell = ws.cell(row=start_row, column=column)
            cell.alignment = Alignment(horizontal='left',
                                       vertical='center')  # Left horizontal alignment and center vertical alignment

    # Merge cells with the same driver name and other columns except the excluded ones
    current_driver_name = None
    start_row = None

    for row in range(2, ws.max_row + 1):  # Assuming the first row is the header
        cell_value = ws[f'A{row}'].value  # Replace 'A' with the actual column letter of 'Lái xe'
        if cell_value == current_driver_name:
            continue
        else:
            if current_driver_name is not None and start_row is not None:
                end_row = row - 1
                for col_idx in range(1, ws.max_column + 1):
                    if col_idx not in exclude_columns_indices:
                        merge_column_range(start_row, end_row, col_idx)
            current_driver_name = cell_value
            start_row = row

    # Handle the last set of merged cells
    if current_driver_name is not None and start_row is not None:
        end_row = ws.max_row
        for col_idx in range(1, ws.max_column + 1):
            if col_idx not in exclude_columns_indices:
                merge_column_range(start_row, end_row, col_idx)

    # Apply alignment to all cells
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(horizontal='left', vertical='center')

    # Save the modified Excel file
    wb.save(output_file_path)

    print(f"Data has been processed, sorted, merged, and saved to {output_file_path}")


url = "https://csht.vnptnghean.com.vn/centers/phieusuachuaphuongtienvantaiexport?type=json"
token = "eyJ0eXAiOiJKV1QiLCJhbGciOiJIUzI1NiJ9.eyJ0b2tlbl90eXBlIjoiYWNjZXNzIiwiZXhwIjoxNzIyNDk2ODE5LCJpYXQiOjE3MjIzMjQwMTksImp0aSI6Ijk4MThmODI3NmY5OTRlZjQ4NGY3NjhmZTFkNjIwMmJiIiwidXNlcl9pZCI6NX0.n55AAjfvZoflqsvezG8NujfF6MFB6yNviV3lmVgiNnI"

# Thiết lập header
headers = {
    "Authorization": f"Bearer {token}"
}

# Gửi yêu cầu GET tới API và lấy dữ liệu
response = requests.get(url, headers=headers)
data = response.json()

data = [
    {
        "Lái xe": "Nguyễn Văn A",
        "lx": "Nguyễn Văn A",
        "Biển kiểm soát": "80A-123.45",
        "bks": "80A-123.45",
        "Định mức": 23.0,
        "Ngày": "2024-08-05",
        "Hạng mục thanh toán": "Xăng A95",
        "Đơn giá": 28000.0,
        "Số lượng": 45.0,
        "Thành tiền": 1260000.0
    },
    {
        "Lái xe": "Nguyễn Văn A",
        "lx": "Nguyễn Văn A",
        "Biển kiểm soát": "80A-123.45",
        "bks": "80A-123.45",
        "Định mức": 23.0,
        "Ngày": "2024-07-30",
        "Hạng mục thanh toán": "Xăng A92",
        "Đơn giá": 27000.0,
        "Số lượng": 60.0,
        "Thành tiền": 1620000.0
    },
    {
        "Lái xe": "Nguyễn Văn A",
        "lx": "Nguyễn Văn A",
        "Biển kiểm soát": "80A-123.45",
        "bks": "80A-123.45",
        "Định mức": 23.0,
        "Ngày": "2024-07-25",
        "Hạng mục thanh toán": "Sửa chữa xe",
        "Đơn giá": 5500000.0,
        "Số lượng": 1.0,
        "Thành tiền": 5500000.0,
        "Phụ phí": 300000.0
    },
    {
        "Lái xe": "Trần Văn B",
        "lx": "Trần Văn B",
        "Biển kiểm soát": "80B-234.56",
        "bks": "80B-234.56",
        "Định mức": 20.0,
        "Ngày": "2024-07-28",
        "Hạng mục thanh toán": "Sửa chữa xe",
        "Đơn giá": 9500000.0,
        "Số lượng": 1.0,
        "Thành tiền": 9500000.0
    },
    {
        "Lái xe": "Trần Văn B",
        "lx": "Trần Văn B",
        "Biển kiểm soát": "80B-234.56",
        "bks": "80B-234.56",
        "Định mức": 20.0,
        "Ngày": "2024-07-25",
        "Hạng mục thanh toán": "Xăng A95",
        "Đơn giá": 28000.0,
        "Số lượng": 70.0,
        "Thành tiền": 1960000.0
    },
    {
        "Lái xe": "Trần Văn B",
        "lx": "Trần Văn B",
        "Biển kiểm soát": "80B-234.56",
        "bks": "80B-234.56",
        "Định mức": 20.0,
        "Ngày": "2024-07-22",
        "Hạng mục thanh toán": "Xăng A92",
        "Đơn giá": 27000.0,
        "Số lượng": 80.0,
        "Thành tiền": 2160000.0
    },
    {
        "Lái xe": "Lê Thị C",
        "lx": "Lê Thị C",
        "Biển kiểm soát": "80C-345.67",
        "bks": "80C-345.67",
        "Định mức": 25.0,
        "Ngày": "2024-07-20",
        "Hạng mục thanh toán": "Sửa chữa xe",
        "Đơn giá": 9000000.0,
        "Số lượng": 1.0,
        "Thành tiền": 9000000.0
    },
    {
        "Lái xe": "Lê Thị C",
        "lx": "Lê Thị C",
        "Biển kiểm soát": "80C-345.67",
        "bks": "80C-345.67",
        "Định mức": 25.0,
        "Ngày": "2024-07-18",
        "Hạng mục thanh toán": "Xăng A95",
        "Đơn giá": 28000.0,
        "Số lượng": 55.0,
        "Thành tiền": 1540000.0,
        "Phí dịch vụ": 100000.0
    },
    {
        "Lái xe": "Lê Thị C",
        "lx": "Lê Thị C",
        "Biển kiểm soát": "80C-345.67",
        "bks": "80C-345.67",
        "Định mức": 25.0,
        "Ngày": "2024-07-15",
        "Hạng mục thanh toán": "Xăng A92",
        "Đơn giá": 27000.0,
        "Số lượng": 100.0,
        "Thành tiền": 2700000.0
    },
    {
        "Lái xe": "Phạm Văn D",
        "lx": "Phạm Văn D",
        "Biển kiểm soát": "80D-456.78",
        "bks": "80D-456.78",
        "Định mức": 22.0,
        "Ngày": "2024-07-29",
        "Hạng mục thanh toán": "Xăng A95",
        "Đơn giá": 28000.0,
        "Số lượng": 60.0,
        "Thành tiền": 1680000.0
    },
    {
        "Lái xe": "Phạm Văn D",
        "lx": "Phạm Văn D",
        "Biển kiểm soát": "80D-456.78",
        "bks": "80D-456.78",
        "Định mức": 22.0,
        "Ngày": "2024-07-25",
        "Hạng mục thanh toán": "Xăng A92",
        "Đơn giá": 27000.0,
        "Số lượng": 75.0,
        "Thành tiền": 2025000.0
    },
    {
        "Lái xe": "Phạm Văn D",
        "lx": "Phạm Văn D",
        "Biển kiểm soát": "80D-456.78",
        "bks": "80D-456.78",
        "Định mức": 22.0,
        "Ngày": "2024-07-22",
        "Hạng mục thanh toán": "Sửa chữa xe",
        "Đơn giá": 7200000.0,
        "Số lượng": 1.0,
        "Thành tiền": 7200000.0,
        "Chi tiết": "Sửa phanh"
    }
]

data_to_excel(data)
